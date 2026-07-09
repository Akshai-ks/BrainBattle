import random
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.contrib.auth.hashers import make_password, check_password
from django.db.models import Avg, Max, Count

from .models import Subject, Student, Game, MCQQuestion, FillBlankQuestion, MatchItem, GameAttempt, AnonymousFeedback, PDFNote, StudentMark, TeacherSetting, Announcement
from .forms import TeacherRegisterForm, StudentRegisterForm, StudentLoginForm

# Custom decorators/checks for student
def student_login_required(view_func):
    def wrapper(request, *args, **kwargs):
        # Allow teachers in trial mode to bypass student login check
        if (request.session.get('trial_mode') == True or request.GET.get('trial') == 'true') and request.user.is_authenticated:
            return view_func(request, *args, **kwargs)
        if not request.session.get('student_id'):
            messages.error(request, "Please login as a student to access this page.")
            return redirect('student_login')
        return view_func(request, *args, **kwargs)
    return wrapper

# General Views
def index(request):
    if request.user.is_authenticated:
        return redirect('teacher_dashboard')
    elif request.session.get('student_id'):
        return redirect('student_dashboard')
    return render(request, 'game_core/index.html')

def logout_view(request):
    if request.user.is_authenticated:
        logout(request)
        messages.success(request, "Teacher logged out successfully.")
    elif request.session.get('student_id'):
        request.session.flush()
        messages.success(request, "Student logged out successfully.")
    return redirect('index')

# Teacher Authentication
def teacher_login(request):
    # Ensure the single teacher superuser exists and has the correct password/permissions
    try:
        user = User.objects.get(username='teacher')
        if not user.check_password('password@teacher') or not user.is_superuser:
            user.set_password('password@teacher')
            user.is_superuser = True
            user.is_staff = True
            user.save()
    except User.DoesNotExist:
        User.objects.create_superuser('teacher', 'teacher@example.com', 'password@teacher')

    if request.user.is_authenticated:
        return redirect('teacher_dashboard')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, f"Welcome back, {user.username}!")
            return redirect('teacher_dashboard')
        else:
            messages.error(request, "Invalid username or password. Please try again.")
    return render(request, 'game_core/teacher_login.html')

# Student Authentication
def student_register(request):
    messages.error(request, "Open student registration is disabled. Please contact your teacher to obtain your credentials.")
    return redirect('student_login')

def student_login(request):
    if request.session.get('student_id'):
        return redirect('student_dashboard')
    if request.method == 'POST':
        form = StudentLoginForm(request.POST)
        if form.is_valid():
            reg_num = form.cleaned_data['register_number']
            password = form.cleaned_data['password']
            try:
                from django.contrib.auth.hashers import check_password
                student = Student.objects.get(register_number=reg_num)
                if student.password and check_password(password, student.password):
                    request.session['student_id'] = student.id
                    request.session['student_name'] = student.name
                    messages.success(request, f"Welcome back, {student.name}!")
                    return redirect('student_dashboard')
                else:
                    messages.error(request, "Invalid registration number or password.")
            except Student.DoesNotExist:
                messages.error(request, "Invalid registration number or password.")
        else:
            messages.error(request, "Please enter valid login details.")
    else:
        form = StudentLoginForm()
    return render(request, 'game_core/student_login.html', {'form': form})

@login_required
def teacher_dashboard(request):
    # Get or create teacher settings
    teacher_settings, _ = TeacherSetting.objects.get_or_create(teacher=request.user)

    context = {
        'teacher_settings': teacher_settings,
    }
    return render(request, 'game_core/teacher_dashboard.html', context)

# Student Dashboard & Game Selection
@student_login_required
def student_dashboard(request):
    student = get_object_or_404(Student, id=request.session['student_id'])
    
    # Load settings for the teacher
    teacher = User.objects.filter(username='teacher').first()
    teacher_settings = None
    if teacher:
        teacher_settings, _ = TeacherSetting.objects.get_or_create(teacher=teacher)
    else:
        first_setting = TeacherSetting.objects.first()
        if first_setting:
            teacher_settings = first_setting
        else:
            default_teacher = User.objects.filter(is_superuser=True).first()
            if default_teacher:
                teacher_settings, _ = TeacherSetting.objects.get_or_create(teacher=default_teacher)
            else:
                teacher_settings = TeacherSetting()
                
    # Fetch Announcements
    announcements = Announcement.objects.all().order_by('-created_at')
    
    # Query games created by this student
    my_created_games = Game.objects.filter(created_by_student=student).order_by('-created_at')
    
    context = {
        'student': student,
        'teacher_settings': teacher_settings,
        'announcements': announcements,
        'my_created_games': my_created_games,
    }
    return render(request, 'game_core/student_dashboard.html', context)


@login_required
def create_game_setup(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        game_type = request.POST.get('game_type')
        instructions = request.POST.get('instructions')
        explanation = request.POST.get('explanation')

        subject, _ = Subject.objects.get_or_create(name="General")

        game = Game.objects.create(
            title=title,
            subject=subject,
            game_type=game_type,
            instructions=instructions,
            explanation=explanation,
            created_by=request.user
        )
        return redirect('setup_game_questions', game_id=game.id)

    return render(request, 'game_core/create_game.html')


@login_required
def setup_game_questions(request, game_id):
    game = get_object_or_404(Game, id=game_id, created_by=request.user)
    
    if request.method == 'POST':
        if game.game_type == 'word_puzzle':
            from .models import WordPuzzleQuestion
            game.word_puzzles.all().delete()
            words = request.POST.getlist('correct_word')
            overrides = request.POST.getlist('letters_override')
            c1s = request.POST.getlist('clue1')
            c1ts = request.POST.getlist('clue1_time')
            c2s = request.POST.getlist('clue2')
            c2ts = request.POST.getlist('clue2_time')
            c3s = request.POST.getlist('clue3')
            fms = request.POST.getlist('full_marks')
            r1s = request.POST.getlist('reduced_marks_clue1')
            m2s = request.POST.getlist('min_marks_clue2')

            for i in range(len(words)):
                if words[i].strip():
                    WordPuzzleQuestion.objects.create(
                        game=game,
                        correct_word=words[i].strip().upper(),
                        letters_override=overrides[i].strip().upper() if i < len(overrides) else '',
                        clue1=c1s[i],
                        clue1_time=int(c1ts[i] or 15),
                        clue2=c2s[i],
                        clue2_time=int(c2ts[i] or 30),
                        clue3=c3s[i] if i < len(c3s) else '',
                        full_marks=int(fms[i] or 10),
                        reduced_marks_clue1=int(r1s[i] or 7),
                        min_marks_clue2=int(m2s[i] or 4)
                    )
            messages.success(request, f"Word Puzzle Game '{game.title}' created successfully!")
            return redirect('teacher_dashboard')

        elif game.game_type == 'mcq':
            # Support multiple questions. Let's delete existing questions and recreate
            game.mcq_questions.all().delete()
            q_texts = request.POST.getlist('q_text')
            opt_as = request.POST.getlist('opt_a')
            opt_bs = request.POST.getlist('opt_b')
            opt_cs = request.POST.getlist('opt_c')
            opt_ds = request.POST.getlist('opt_d')
            corrects = request.POST.getlist('correct')
            c1s = request.POST.getlist('clue1')
            c2s = request.POST.getlist('clue2')
            c3s = request.POST.getlist('clue3')

            for i in range(len(q_texts)):
                if q_texts[i].strip():
                    MCQQuestion.objects.create(
                        game=game,
                        question_text=q_texts[i],
                        option_a=opt_as[i],
                        option_b=opt_bs[i],
                        option_c=opt_cs[i],
                        option_d=opt_ds[i],
                        correct_option=corrects[i],
                        clue1=c1s[i] if i < len(c1s) else '',
                        clue2=c2s[i] if i < len(c2s) else '',
                        clue3=c3s[i] if i < len(c3s) else ''
                    )
            messages.success(request, f"MCQ Game '{game.title}' created with questions!")
            return redirect('teacher_dashboard')

        elif game.game_type == 'fill_blanks':
            game.blank_questions.all().delete()
            q_texts = request.POST.getlist('q_text')
            answers = request.POST.getlist('correct_answer')
            c1s = request.POST.getlist('clue1')
            c2s = request.POST.getlist('clue2')
            c3s = request.POST.getlist('clue3')

            for i in range(len(q_texts)):
                if q_texts[i].strip():
                    FillBlankQuestion.objects.create(
                        game=game,
                        question_text=q_texts[i],
                        correct_answer=answers[i].strip(),
                        clue=c1s[i].strip() if i < len(c1s) else '',
                        clue1=c1s[i].strip() if i < len(c1s) else '',
                        clue2=c2s[i].strip() if i < len(c2s) else '',
                        clue3=c3s[i].strip() if i < len(c3s) else ''
                    )
            messages.success(request, f"Fill in the Blanks Game '{game.title}' created successfully!")
            return redirect('teacher_dashboard')

        elif game.game_type == 'match_following':
            game.match_items.all().delete()
            lefts = request.POST.getlist('left_item')
            rights = request.POST.getlist('right_item')

            game.clue1 = request.POST.get('clue1', '').strip()
            game.clue2 = request.POST.get('clue2', '').strip()
            game.clue3 = request.POST.get('clue3', '').strip()
            game.save()

            for i in range(len(lefts)):
                if lefts[i].strip() and rights[i].strip():
                    MatchItem.objects.create(
                        game=game,
                        left_item=lefts[i].strip(),
                        right_item=rights[i].strip()
                    )
            messages.success(request, f"Match the Following Game '{game.title}' created successfully!")
            return redirect('teacher_dashboard')

    return render(request, f'game_core/setup_{game.game_type}.html', {'game': game})

@login_required
def delete_game(request, game_id):
    from django.db.models import Q
    game = get_object_or_404(Game, Q(id=game_id) & (Q(created_by=request.user) | Q(created_by_student__isnull=False)))
    game.delete()
    messages.success(request, "Game deleted successfully.")
    return redirect('view_games')

@login_required
def toggle_game_active(request, game_id):
    from django.db.models import Q
    game = get_object_or_404(Game, Q(id=game_id) & (Q(created_by=request.user) | Q(created_by_student__isnull=False)))
    game.is_active = not game.is_active
    game.save()
    status = "activated" if game.is_active else "deactivated"
    messages.success(request, f"Game '{game.title}' has been {status}.")
    return redirect('view_games')

def get_game_leaderboard(game, limit=10):
    from .models import GameAttempt
    # Fetch all attempts for this game ordered by score DESC, time_taken ASC, and completed_at ASC
    attempts = GameAttempt.objects.filter(game=game).order_by('-score', 'time_taken', 'completed_at')
    seen_students = set()
    leaderboard = []
    for att in attempts:
        if att.student_id not in seen_students:
            seen_students.add(att.student_id)
            leaderboard.append(att)
            if limit and len(leaderboard) >= limit:
                break
    return leaderboard

# Gameplay Views
@student_login_required

def play_game(request, game_id):
    # Determine if this request triggers trial mode
    is_trial_request = request.GET.get('trial') == 'true' and request.user.is_authenticated
    if is_trial_request:
        request.session['trial_mode'] = True
    else:
        if not request.user.is_authenticated:
            request.session.pop('trial_mode', None)

    trial_active = request.session.get('trial_mode') == True and request.user.is_authenticated
    
    game = get_object_or_404(Game, id=game_id)
    
    if not trial_active:
        if not game.is_active:
            messages.error(request, "This game is not currently active.")
            return redirect('student_dashboard')
            
        student_id = request.session.get('student_id')
        student = get_object_or_404(Student, id=student_id)
        
        # Create an entry record to track student starting the game
        from .models import GameEntry
        GameEntry.objects.create(student=student, game=game)
        
        # Load teacher settings

        teacher = User.objects.filter(username='teacher').first()
        teacher_settings = None
        if teacher:
            teacher_settings, _ = TeacherSetting.objects.get_or_create(teacher=teacher)
        else:
            teacher_settings = TeacherSetting()
            
        if not teacher_settings.games_enabled:
            messages.error(request, "Conceptual learning games are currently disabled by the teacher.")
            return redirect('student_dashboard')
            
        # Check block state for this student and game
        from .models import StudentGamePlayPermission, GameAttempt
        attempts_count = GameAttempt.objects.filter(student=student, game=game).count()
        permission = StudentGamePlayPermission.objects.filter(student=student, game=game).first()
        allowed = permission.allowed_attempts if permission else 1
        
        # Block if replays disabled globally and they played at least once
        if not teacher_settings.replays_enabled and attempts_count >= 1:
            messages.error(request, f"Replay has been disabled globally by the teacher. You cannot play '{game.title}' again.")
            return redirect('student_dashboard')
            
        if attempts_count >= allowed:
            messages.error(request, f"You have already completed '{game.title}'. Replays are locked unless unlocked by your teacher.")
            return redirect('student_dashboard')
            
    if game.game_type == 'word_puzzle':
        puzzles = list(game.word_puzzles.all())
        random.shuffle(puzzles)
        
        prepared_puzzles = []
        total_max_score = 0
        
        for p in puzzles:
            word = p.correct_word.upper()
            total_max_score += p.full_marks
            
            # Prepare letters list
            letters = list(word)
            if p.letters_override:
                letters = [l.strip().upper() for l in p.letters_override.split(',') if l.strip()]
            else:
                fillers = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
                random.shuffle(fillers)
                while len(letters) < 14:
                    filler = fillers.pop()
                    if filler not in letters or random.random() > 0.5:
                        letters.append(filler)
            
            random.shuffle(letters)
            
            prepared_puzzles.append({
                'id': p.id,
                'correct_word': word,
                'letters': letters,
                'clue1': p.clue1,
                'clue1_time': p.clue1_time,
                'clue2': p.clue2,
                'clue2_time': p.clue2_time,
                'clue3': p.clue3,
                'full_marks': p.full_marks,
                'reduced_marks_clue1': p.reduced_marks_clue1,
                'min_marks_clue2': p.min_marks_clue2
            })
            
        import json
        context = {
            'game': game,
            'puzzles_json': json.dumps(prepared_puzzles),
            'total_questions': len(prepared_puzzles),
            'total_max_score': total_max_score,
            'trial_mode': trial_active,
        }
        return render(request, 'game_core/play_word_puzzle.html', context)

    elif game.game_type == 'mcq':
        questions = list(game.mcq_questions.all())
        # Randomize question order for gameplay
        random.shuffle(questions)
        context = {
            'game': game,
            'questions': questions,
            'total_questions': len(questions),
            'trial_mode': trial_active,
        }
        return render(request, 'game_core/play_mcq.html', context)

    elif game.game_type == 'fill_blanks':
        questions = list(game.blank_questions.all())
        random.shuffle(questions)
        context = {
            'game': game,
            'questions': questions,
            'total_questions': len(questions),
            'trial_mode': trial_active,
        }
        return render(request, 'game_core/play_fill_blanks.html', context)

    elif game.game_type == 'match_following':
        items = list(game.match_items.all())
        left_items = [item.left_item for item in items]
        right_items = [item.right_item for item in items]
        
        # Shuffle both
        random.shuffle(left_items)
        random.shuffle(right_items)
        
        context = {
            'game': game,
            'left_items': left_items,
            'right_items': right_items,
            'pairs_json': {item.left_item: item.right_item for item in items},
            'trial_mode': trial_active,
        }
        return render(request, 'game_core/play_match_following.html', context)

    raise Http404("Game type not found.")

@student_login_required
def submit_score(request, game_id):
    if request.method == 'POST':
        game = get_object_or_404(Game, id=game_id)
        
        trial_active = request.session.get('trial_mode') == True and request.user.is_authenticated
        if trial_active:
            request.session['trial_game_id'] = game.id
            request.session['trial_score'] = int(request.POST.get('score', 0))
            request.session['trial_max_score'] = int(request.POST.get('max_score', 10))
            request.session['trial_clues_used'] = int(request.POST.get('clues_used', 0))
            request.session['trial_time_taken'] = int(request.POST.get('time_taken', 0))
            return JsonResponse({'status': 'success', 'attempt_id': 0})
            
        student = get_object_or_404(Student, id=request.session['student_id'])
        
        score = int(request.POST.get('score', 0))
        max_score = int(request.POST.get('max_score', 10))
        clues_used = int(request.POST.get('clues_used', 0))
        time_taken = int(request.POST.get('time_taken', 0))
        
        attempt = GameAttempt.objects.create(
            student=student,
            game=game,
            score=score,
            max_score=max_score,
            clues_used=clues_used,
            time_taken=time_taken
        )
        return JsonResponse({'status': 'success', 'attempt_id': attempt.id})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=400)

# Results & Certificates
@student_login_required
def game_result(request, attempt_id):
    trial_active = request.session.get('trial_mode') == True and request.user.is_authenticated
    
    if trial_active and attempt_id == 0:
        class MockAttempt:
            def __init__(self):
                self.id = 0
                self.score = 0
                self.max_score = 10
                self.clues_used = 0
                self.time_taken = 0
                self.student = None
                self.game = None
                
        attempt = MockAttempt()
        attempt.score = request.session.get('trial_score', 0)
        attempt.max_score = request.session.get('trial_max_score', 10)
        attempt.clues_used = request.session.get('trial_clues_used', 0)
        attempt.time_taken = request.session.get('trial_time_taken', 0)
        attempt.student = type('MockStudent', (object,), {'name': 'Teacher (Trial Mode)'})()
        
        game_id = request.session.get('trial_game_id')
        attempt.game = get_object_or_404(Game, id=game_id) if game_id else None
    else:
        attempt = get_object_or_404(GameAttempt, id=attempt_id, student_id=request.session['student_id'])
    
    # Calculate performance feedback message
    percentage = (attempt.score / attempt.max_score) * 100 if attempt.max_score > 0 else 0
    if percentage >= 100:
        feedback = "Outstanding! Perfect Score! 🏆"
        eligible_for_certificate = True
    elif percentage >= 75:
        feedback = "Great job! Almost perfect! 🌟"
        eligible_for_certificate = False
    elif percentage >= 50:
        feedback = "Good effort! Keep practicing! 👍"
        eligible_for_certificate = False
    else:
        feedback = "Keep learning and try again! 📚"
        eligible_for_certificate = False

    if trial_active:
        eligible_for_certificate = False

    leaderboard = []
    if attempt.game:
        leaderboard = get_game_leaderboard(attempt.game, limit=10)

    context = {
        'attempt': attempt,
        'percentage': round(percentage, 1),
        'feedback': feedback,
        'eligible_for_certificate': eligible_for_certificate,
        'trial_mode': trial_active,
        'leaderboard': leaderboard,
    }
    return render(request, 'game_core/game_result.html', context)


@student_login_required
def generate_certificate(request, attempt_id):
    if request.session.get('trial_mode') == True:
        messages.error(request, "Certificates cannot be generated in Trial Mode.")
        return redirect('view_games')
        
    attempt = get_object_or_404(GameAttempt, id=attempt_id, student_id=request.session['student_id'])
    if attempt.score < attempt.max_score:
        messages.error(request, "Certificates are only generated for a perfect score.")
        return redirect('game_result', attempt_id=attempt_id)
        
    return render(request, 'game_core/certificate.html', {'attempt': attempt})

@student_login_required
def student_game_leaderboard(request, game_id):
    game = get_object_or_404(Game, id=game_id)
    
    teacher = User.objects.filter(username='teacher').first()
    teacher_settings = None
    if teacher:
        teacher_settings, _ = TeacherSetting.objects.get_or_create(teacher=teacher)
    else:
        teacher_settings = TeacherSetting()
        
    if not teacher_settings.games_enabled:
        messages.error(request, "Conceptual learning games are currently disabled by the teacher.")
        return redirect('student_dashboard')
        
    leaderboard = get_game_leaderboard(game, limit=10)
    
    context = {
        'game': game,
        'leaderboard': leaderboard,
    }
    return render(request, 'game_core/student_game_leaderboard.html', context)


@student_login_required
def submit_feedback(request):
    if request.method == 'POST':
        message = request.POST.get('message')
        
        if message and message.strip():
            teachers = User.objects.all()
            if teachers.exists():
                for teacher in teachers:
                    AnonymousFeedback.objects.create(
                        teacher=teacher,
                        message=message.strip()
                    )
                messages.success(request, "Your anonymous feedback has been sent successfully!")
            else:
                messages.error(request, "No teachers registered in the system.")
        else:
            messages.error(request, "Please write your message.")
            
    return redirect('student_dashboard')

@login_required
def download_feedback_pdf(request):
    feedback_list = request.user.received_feedback.all().order_by('-created_at')
    
    # Generate custom PDF byte stream
    title = f"Anonymous Student Feedback for {request.user.username}"
    
    # We will format the text into lines first. Let's wrap long lines to fit in a page.
    def wrap_text(text, max_chars=80):
        # simple word wrap
        words = text.split()
        lines = []
        current_line = []
        current_length = 0
        for w in words:
            if current_length + len(w) + 1 > max_chars:
                lines.append(" ".join(current_line))
                current_line = [w]
                current_length = len(w)
            else:
                current_line.append(w)
                current_length += len(w) + 1
        if current_line:
            lines.append(" ".join(current_line))
        return lines

    def make_latin1_safe(text):
        replacements = {
            '\u2018': "'",
            '\u2019': "'",
            '\u201c': '"',
            '\u201d': '"',
            '\u2013': '-',
            '\u2014': '-',
        }
        for orig, repl in replacements.items():
            text = text.replace(orig, repl)
        return text.encode('latin-1', errors='replace').decode('latin-1')

    all_lines = []
    for idx, msg in enumerate(feedback_list):
        date_str = msg.created_at.strftime("%Y-%m-%d %H:%M UTC")
        all_lines.append(f"--- Suggestion #{idx + 1} ({date_str}) ---")
        safe_msg = make_latin1_safe(msg.message)
        wrapped = wrap_text(safe_msg, max_chars=75)
        all_lines.extend(wrapped)
        all_lines.append("") # empty line after message

    # Paginate
    lines_per_page = 35
    pages_data = []
    for i in range(0, len(all_lines), lines_per_page):
        pages_data.append(all_lines[i:i + lines_per_page])
        
    if not pages_data:
        pages_data = [[f"No anonymous feedback received yet."]]

    num_pages = len(pages_data)
    page_obj_ids = [4 + i for i in range(num_pages)]
    content_obj_ids = [4 + num_pages + i for i in range(num_pages)]
    font_obj_id = 4 + 2 * num_pages
    
    objects = {}
    objects[1] = b"<< /Type /Catalog /Pages 2 0 R >>"
    
    kids_str = " ".join([f"{pid} 0 R" for pid in page_obj_ids])
    objects[2] = f"<< /Type /Pages /Kids [{kids_str}] /Count {num_pages} >>".encode('latin1')
    objects[3] = b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>"
    
    for i in range(num_pages):
        page_id = page_obj_ids[i]
        content_id = content_obj_ids[i]
        page_lines = pages_data[i]
        
        objects[page_id] = f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Contents {content_id} 0 R /Resources << /Font << /F1 3 0 R >> >> >>".encode('latin1')
        
        stream_lines = [
            b"BT",
            b"/F1 16 Tf",
            b"50 800 Td",
        ]
        
        if i == 0:
            stream_lines.append(f"({title}) Tj".encode('latin1'))
            stream_lines.append(b"0 -30 Td")
            stream_lines.append(b"/F1 12 Tf")
        else:
            stream_lines.append(f"({title} - Page {i+1}) Tj".encode('latin1'))
            stream_lines.append(b"0 -30 Td")
            stream_lines.append(b"/F1 12 Tf")
            
        for line in page_lines:
            # Escape parenthesis and backslashes in PDF string
            escaped = line.replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')
            stream_lines.append(f"0 -20 Td ({escaped}) Tj".encode('latin1'))
            
        stream_lines.append(b"ET")
        stream_data = b"\n".join(stream_lines)
        
        objects[content_id] = f"<< /Length {len(stream_data)} >>\nstream\n".encode('latin1') + stream_data + b"\nendstream"
        
    body = b"%PDF-1.4\n"
    offsets = {}
    for num in sorted(objects.keys()):
        offsets[num] = len(body)
        body += f"{num} 0 obj\n".encode('latin1')
        body += objects[num] + b"\nendobj\n"
        
    xref_start = len(body)
    body += b"xref\n"
    body += f"0 {len(objects) + 1}\n".encode('latin1')
    body += b"0000000000 65535 f \n"
    for num in sorted(objects.keys()):
        body += f"{offsets[num]:010d} 00000 n \n".encode('latin1')
        
    body += b"trailer\n"
    body += f"<< /Size {len(objects) + 1} /Root 1 0 R >>\n".encode('latin1')
    body += b"startxref\n"
    body += f"{xref_start}\n".encode('latin1')
    body += b"%%EOF\n"
    
    from django.http import HttpResponse
    response = HttpResponse(body, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="student_suggestions.pdf"'
    return response

@login_required
def delete_all_feedback(request):
    if request.method == 'POST':
        count = request.user.received_feedback.all().count()
        request.user.received_feedback.all().delete()
        messages.success(request, f"Successfully deleted all {count} feedback messages.")
    return redirect('manage_announcements')

@login_required
def manage_students(request):
    from django.contrib.auth.hashers import make_password
    import openpyxl
    
    students = Student.objects.all().order_by('-created_at')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'manual':
            name = request.POST.get('name')
            register_number = request.POST.get('register_number')
            
            if not name or not register_number:
                messages.error(request, "Please provide both Name and Register Number.")
            else:
                name = name.strip()
                register_number = register_number.strip()
                
                if Student.objects.filter(register_number=register_number).exists():
                    messages.error(request, f"A student with Register Number '{register_number}' already exists.")
                else:
                    default_password = f"{register_number}@{register_number}"
                    Student.objects.create(
                        name=name,
                        register_number=register_number,
                        password=make_password(default_password)
                    )
                    messages.success(request, f"Student '{name}' created successfully with password '{default_password}'.")
                    return redirect('manage_students')
                    
        elif action == 'excel':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, "Please select an Excel file to upload.")
            else:
                try:
                    wb = openpyxl.load_workbook(excel_file, read_only=True)
                    sheet = wb.active
                    
                    success_count = 0
                    skip_count = 0
                    
                    header_row = None
                    for row in sheet.iter_rows(values_only=True):
                        if any(row):  
                            header_row = row
                            break
                    
                    name_idx = 0
                    reg_idx = 1
                    
                    if header_row:
                        header_lower = [str(cell).lower().strip() if cell is not None else "" for cell in header_row]
                        for idx, val in enumerate(header_lower):
                            if "name" in val:
                                name_idx = idx
                            elif "register" in val or "reg" in val or "number" in val or "roll" in val:
                                reg_idx = idx
                                
                    first_row_skipped = False
                    for row in sheet.iter_rows(values_only=True):
                        if not any(row):
                            continue
                        
                        if not first_row_skipped and header_row and row == header_row:
                            first_row_skipped = True
                            continue
                            
                        name = str(row[name_idx]).strip() if len(row) > name_idx and row[name_idx] is not None else None
                        register_number = str(row[reg_idx]).strip() if len(row) > reg_idx and row[reg_idx] is not None else None
                        
                        if not name or not register_number or name.lower() == 'none' or register_number.lower() == 'none':
                            continue
                        
                        if Student.objects.filter(register_number=register_number).exists():
                            skip_count += 1
                        else:
                            default_password = f"{register_number}@{register_number}"
                            Student.objects.create(
                                name=name,
                                register_number=register_number,
                                password=make_password(default_password)
                            )
                            success_count += 1
                            
                    messages.success(request, f"Imported {success_count} students successfully. Skipped {skip_count} duplicates/invalid rows.")
                    return redirect('manage_students')
                except Exception as e:
                    messages.error(request, f"Failed to process Excel file: {str(e)}")
                    
    context = {
        'students': students
    }
    return render(request, 'game_core/manage_students.html', context)

@login_required
def delete_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    name = student.name
    student.delete()
    messages.success(request, f"Student '{name}' deleted successfully.")
    return redirect('view_students')

@login_required
def delete_all_students(request):
    if request.method == 'POST':
        count = Student.objects.count()
        Student.objects.all().delete()
        messages.success(request, f"Successfully deleted all {count} student records.")
    return redirect('view_students')

@login_required
def unlock_student_replay(request, student_id, game_id):
    from .models import Student, Game, StudentGamePlayPermission, GameAttempt
    student = get_object_or_404(Student, id=student_id)
    game = get_object_or_404(Game, id=game_id)
    
    attempts_count = GameAttempt.objects.filter(student=student, game=game).count()
    permission, created = StudentGamePlayPermission.objects.get_or_create(student=student, game=game)
    
    permission.allowed_attempts = max(permission.allowed_attempts, attempts_count) + 1
    permission.save()
    
    messages.success(request, f"Replay access unlocked for student '{student.name}' on game '{game.title}'.")
    return redirect('teacher_dashboard')

@login_required
def unlock_all_student_replays(request):
    from .models import StudentGamePlayPermission, GameAttempt
    attempts = GameAttempt.objects.all()
    for att in attempts:
        attempts_count = GameAttempt.objects.filter(student=att.student, game=att.game).count()
        perm, created = StudentGamePlayPermission.objects.get_or_create(student=att.student, game=att.game)
        perm.allowed_attempts = max(perm.allowed_attempts, attempts_count) + 1
        perm.save()
    messages.success(request, "Unlocked replay access for all students across all games!")
    return redirect('teacher_dashboard')

@login_required
def lock_all_student_replays(request):
    from .models import StudentGamePlayPermission, GameAttempt
    attempts = GameAttempt.objects.all()
    for att in attempts:
        attempts_count = GameAttempt.objects.filter(student=att.student, game=att.game).count()
        perm, created = StudentGamePlayPermission.objects.get_or_create(student=att.student, game=att.game)
        perm.allowed_attempts = attempts_count
        perm.save()
    messages.success(request, "Blocked replay access for all completed games for all students!")
    return redirect('teacher_dashboard')

# PDF Notes Upload & Management
@login_required
def manage_notes(request):
    notes = PDFNote.objects.all().select_related('subject', 'uploaded_by').order_by('-uploaded_at')
    subjects = Subject.objects.all()
    
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        subject_id = request.POST.get('subject_id')
        new_subject_name = request.POST.get('new_subject', '').strip()
        file = request.FILES.get('file')
        
        if not title:
            messages.error(request, "Please provide a note title.")
        elif not file:
            messages.error(request, "Please select a PDF file to upload.")
        elif not file.name.endswith('.pdf'):
            messages.error(request, "Only PDF files are supported.")
        else:
            subject = None
            if new_subject_name:
                subject, _ = Subject.objects.get_or_create(name=new_subject_name)
            elif subject_id:
                subject = get_object_or_404(Subject, id=subject_id)
            
            if not subject:
                subject, _ = Subject.objects.get_or_create(name="General")
                
            PDFNote.objects.create(
                title=title,
                description=description,
                subject=subject,
                file=file,
                uploaded_by=request.user
            )
            messages.success(request, f"Note '{title}' uploaded successfully.")
            return redirect('manage_notes')
            
    return render(request, 'game_core/manage_notes.html', {'notes': notes, 'subjects': subjects})

@login_required
def edit_note(request, note_id):
    note = get_object_or_404(PDFNote, id=note_id)
    subjects = Subject.objects.all()
    
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        subject_id = request.POST.get('subject_id')
        new_subject_name = request.POST.get('new_subject', '').strip()
        file = request.FILES.get('file')
        
        if not title:
            messages.error(request, "Please provide a note title.")
        else:
            subject = None
            if new_subject_name:
                subject, _ = Subject.objects.get_or_create(name=new_subject_name)
            elif subject_id:
                subject = get_object_or_404(Subject, id=subject_id)
            
            if not subject:
                subject = note.subject
                
            note.title = title
            note.description = description
            note.subject = subject
            if file:
                if file.name.endswith('.pdf'):
                    if note.file:
                        note.file.delete(save=False)
                    note.file = file
                else:
                    messages.error(request, "Uploaded file was not a PDF, keeping existing file.")
            note.save()
            messages.success(request, f"Note '{title}' updated successfully.")
            return redirect('manage_notes')
            
    return render(request, 'game_core/edit_note.html', {'note': note, 'subjects': subjects})

@login_required
def delete_note(request, note_id):
    note = get_object_or_404(PDFNote, id=note_id)
    title = note.title
    if note.file:
        note.file.delete(save=False)
    note.delete()
    messages.success(request, f"Note '{title}' deleted successfully.")
    return redirect('manage_notes')

# Student Marks Management
@login_required
def manage_marks(request):
    import openpyxl
    marks = StudentMark.objects.all().select_related('student', 'subject').order_by('-created_at')
    for m in marks:
        m.percentage = (m.marks / m.max_marks) * 100 if m.max_marks > 0 else 0
        
    subjects = Subject.objects.all()
    students = Student.objects.all()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'manual':
            register_number = request.POST.get('register_number', '').strip()
            student_name = request.POST.get('student_name', '').strip()
            subject_id = request.POST.get('subject_id')
            new_subject_name = request.POST.get('new_subject', '').strip()
            marks_val_str = request.POST.get('marks', '').strip()
            max_marks_str = request.POST.get('max_marks', '').strip()
            exam_type = request.POST.get('exam_type', '').strip()
            
            if not register_number or not marks_val_str or not exam_type:
                messages.error(request, "Please provide Register Number, Marks, and Exam Type.")
            else:
                try:
                    marks_val = int(marks_val_str)
                    max_marks_val = int(max_marks_str) if max_marks_str else 100
                    
                    if max_marks_val <= 0:
                        messages.error(request, "Maximum marks must be a positive integer.")
                    elif marks_val > max_marks_val:
                        messages.error(request, f"Marks obtained ({marks_val}) cannot exceed maximum marks ({max_marks_val}).")
                    else:
                        student = None
                        try:
                            student = Student.objects.get(register_number=register_number)
                        except Student.DoesNotExist:
                            name = student_name if student_name else f"Student {register_number}"
                            default_password = f"{register_number}@{register_number}"
                            student = Student.objects.create(
                                name=name,
                                register_number=register_number,
                                password=make_password(default_password)
                            )
                            messages.info(request, f"Registered new student '{name}' with password '{default_password}'.")
                        
                        subject = None
                        if new_subject_name:
                            subject, _ = Subject.objects.get_or_create(name=new_subject_name)
                        elif subject_id:
                            subject = get_object_or_404(Subject, id=subject_id)
                            
                        if not subject:
                            subject, _ = Subject.objects.get_or_create(name="General")
                        
                        StudentMark.objects.update_or_create(
                            student=student,
                            subject=subject,
                            exam_type=exam_type,
                            defaults={'marks': marks_val, 'max_marks': max_marks_val, 'uploaded_by': request.user}
                        )
                        messages.success(request, f"Marks updated for {student.name} in {subject.name} ({exam_type}).")
                        return redirect('manage_marks')
                except ValueError:
                    messages.error(request, "Marks and Maximum Marks must be integers.")
                    
        elif action == 'excel':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, "Please select an Excel file.")
            else:
                try:
                    wb = openpyxl.load_workbook(excel_file, data_only=True)
                    sheet = wb.active
                    
                    headers = None
                    rows = list(sheet.iter_rows(values_only=True))
                    
                    header_row_idx = 0
                    for idx, r in enumerate(rows):
                        if r and any(r):
                            headers = [str(c).lower().strip() if c is not None else "" for c in r]
                            header_row_idx = idx
                            break
                            
                    if not headers:
                        raise Exception("Could not find header row in Excel file.")
                        
                    reg_idx = -1
                    name_idx = -1
                    sub_idx = -1
                    marks_idx = -1
                    max_marks_idx = -1
                    exam_idx = -1
                    
                    for idx, h in enumerate(headers):
                        if 'register' in h or 'reg' in h or 'roll' in h or 'number' in h:
                            reg_idx = idx
                        elif 'name' in h:
                            name_idx = idx
                        elif 'subject' in h or 'sub' in h:
                            sub_idx = idx
                        elif 'max' in h or 'out of' in h or 'total' in h:
                            max_marks_idx = idx
                        elif 'marks' in h or 'mark' in h or 'score' in h:
                            if max_marks_idx != idx:
                                marks_idx = idx
                        elif 'exam' in h:
                            exam_idx = idx
                            
                    if reg_idx == -1 or sub_idx == -1 or marks_idx == -1 or exam_idx == -1:
                        raise Exception("Excel must contain columns for Register Number, Subject, Marks, and Exam Type (Name and Max Marks are optional).")
                        
                    success_count = 0
                    skip_count = 0
                    
                    for r in rows[header_row_idx + 1:]:
                        if not r or not any(r):
                            continue
                            
                        reg_val = str(r[reg_idx]).strip() if reg_idx < len(r) and r[reg_idx] is not None else None
                        sub_val = str(r[sub_idx]).strip() if sub_idx < len(r) and r[sub_idx] is not None else None
                        marks_val = r[marks_idx] if marks_idx < len(r) else None
                        max_marks_val = r[max_marks_idx] if (max_marks_idx != -1 and max_marks_idx < len(r)) else None
                        exam_val = str(r[exam_idx]).strip() if exam_idx < len(r) and r[exam_idx] is not None else None
                        name_val = str(r[name_idx]).strip() if (name_idx != -1 and name_idx < len(r) and r[name_idx] is not None) else None
                        
                        if not reg_val or reg_val.lower() == 'none' or not sub_val or marks_val is None or not exam_val:
                            skip_count += 1
                            continue
                            
                        try:
                            marks_int = int(float(marks_val))
                            
                            max_marks_int = 100
                            if max_marks_val is not None:
                                try:
                                    max_marks_int = int(float(max_marks_val))
                                except ValueError:
                                    max_marks_int = 100
                                    
                            if marks_int > max_marks_int:
                                # Over limit validation fail
                                skip_count += 1
                                continue
                            
                            student = None
                            try:
                                student = Student.objects.get(register_number=reg_val)
                            except Student.DoesNotExist:
                                student_name = name_val if name_val else f"Student {reg_val}"
                                default_password = f"{reg_val}@{reg_val}"
                                student = Student.objects.create(
                                    name=student_name,
                                    register_number=reg_val,
                                    password=make_password(default_password)
                                )
                                
                            subject, _ = Subject.objects.get_or_create(name=sub_val)
                            
                            StudentMark.objects.update_or_create(
                                student=student,
                                subject=subject,
                                exam_type=exam_val,
                                defaults={'marks': marks_int, 'max_marks': max_marks_int, 'uploaded_by': request.user}
                            )
                            success_count += 1
                        except Exception as row_error:
                            skip_count += 1
                            continue
                            
                    messages.success(request, f"Imported {success_count} student marks. Skipped {skip_count} invalid rows.")
                    return redirect('manage_marks')
                except Exception as e:
                    messages.error(request, f"Failed to parse Excel: {str(e)}")
                    
    return render(request, 'game_core/manage_marks.html', {
        'marks': marks,
        'subjects': subjects,
        'students': students
    })

@login_required
def edit_mark(request, mark_id):
    mark = get_object_or_404(StudentMark, id=mark_id)
    
    if request.method == 'POST':
        marks_str = request.POST.get('marks', '').strip()
        max_marks_str = request.POST.get('max_marks', '').strip()
        exam_type = request.POST.get('exam_type', '').strip()
        
        if not marks_str or not exam_type:
            messages.error(request, "Please provide Marks and Exam Type.")
        else:
            try:
                marks_val = int(marks_str)
                max_marks_val = int(max_marks_str) if max_marks_str else 100
                
                if max_marks_val <= 0:
                    messages.error(request, "Maximum marks must be a positive integer.")
                elif marks_val > max_marks_val:
                    messages.error(request, f"Marks obtained ({marks_val}) cannot exceed maximum marks ({max_marks_val}).")
                else:
                    mark.marks = marks_val
                    mark.max_marks = max_marks_val
                    mark.exam_type = exam_type
                    mark.save()
                    messages.success(request, f"Marks updated successfully for {mark.student.name}.")
                    return redirect('manage_marks')
            except ValueError:
                messages.error(request, "Marks and Maximum Marks must be integers.")
                
    return render(request, 'game_core/edit_mark.html', {'mark': mark})

@login_required
def delete_mark(request, mark_id):
    mark = get_object_or_404(StudentMark, id=mark_id)
    student_name = mark.student.name
    mark.delete()
    messages.success(request, f"Marks record for {student_name} deleted successfully.")
    return redirect('manage_marks')

@login_required
def delete_all_marks(request):
    if request.method == 'POST':
        count = StudentMark.objects.count()
        StudentMark.objects.all().delete()
        messages.success(request, f"Successfully deleted all {count} marks records.")
    return redirect('manage_marks')

@login_required
def toggle_student_replay(request, student_id, game_id):
    from .models import Student, Game, StudentGamePlayPermission, GameAttempt
    student = get_object_or_404(Student, id=student_id)
    game = get_object_or_404(Game, id=game_id)
    
    attempts_count = GameAttempt.objects.filter(student=student, game=game).count()
    permission, created = StudentGamePlayPermission.objects.get_or_create(student=student, game=game)
    
    if permission.allowed_attempts > attempts_count:
        permission.allowed_attempts = attempts_count
        status = "blocked (OFF)"
    else:
        permission.allowed_attempts = max(permission.allowed_attempts, attempts_count) + 1
        status = "playable (ON)"
        
    permission.save()
    messages.success(request, f"Replay status for '{student.name}' on '{game.title}' is now {status}.")
    if request.GET.get('next') == 'view_students':
        return redirect('view_students')
    return redirect('teacher_dashboard')

@login_required
def update_teacher_settings(request):
    if request.method == 'POST':
        teacher_settings, _ = TeacherSetting.objects.get_or_create(teacher=request.user)
        teacher_settings.anonymous_feedback_enabled = 'anonymous_feedback_enabled' in request.POST
        teacher_settings.replays_enabled = 'replays_enabled' in request.POST
        teacher_settings.games_enabled = 'games_enabled' in request.POST
        teacher_settings.save()
        messages.success(request, "Platform configurations updated successfully.")
    return redirect('teacher_dashboard')

# Redesigned Dashboard Subsections
@login_required
def view_games(request):
    from django.db.models import Q, Avg, Max, Min
    from .models import GameEntry, GameAttempt
    all_games = Game.objects.filter(
        Q(created_by=request.user) | Q(created_by_student__isnull=False, approval_status='approved')
    ).order_by('-created_at')
    
    # Enrich all games with their player tracking stats and leaderboard
    for game in all_games:
        game.total_players = GameEntry.objects.filter(game=game).values('student').distinct().count()
        game.total_entries = GameEntry.objects.filter(game=game).count()
        
        attempts_qs = GameAttempt.objects.filter(game=game)
        game.total_attempts = attempts_qs.count()
        
        avg_score = attempts_qs.aggregate(avg=Avg('score'))['avg']
        high_score = attempts_qs.aggregate(max=Max('score'))['max']
        avg_time = attempts_qs.aggregate(avg=Avg('time_taken'))['avg']
        
        game.avg_score = round(avg_score, 1) if avg_score is not None else 0
        game.high_score = high_score if high_score is not None else 0
        game.avg_time = round(avg_time, 1) if avg_time is not None else 0
        
        # Get leaderboard
        game.leaderboard_top5 = get_game_leaderboard(game, limit=5)
        
    active_games = all_games.filter(is_active=True)
    inactive_games = all_games.filter(is_active=False)
    
    return render(request, 'game_core/view_games.html', {
        'active_games': active_games,
        'inactive_games': inactive_games,
    })


@login_required
def view_students(request):
    from django.db.models import Q, Count
    from .models import StudentGamePlayPermission, GameAttempt, Game
    
    query = request.GET.get('search', '').strip()
    students = Student.objects.all().order_by('-created_at')
    
    if query:
        students = students.filter(
            Q(name__icontains=query) | Q(register_number__icontains=query) | Q(email__icontains=query)
        )
        
    games = Game.objects.filter(
        Q(created_by=request.user) | Q(created_by_student__isnull=False, approval_status='approved')
    ).order_by('title')
    
    permissions = StudentGamePlayPermission.objects.all()
    attempts = GameAttempt.objects.all().values('student_id', 'game_id').annotate(count=Count('id'))
    
    perm_map = {}
    for p in permissions:
        perm_map[(p.student_id, p.game_id)] = p.allowed_attempts
        
    att_map = {}
    for a in attempts:
        att_map[(a['student_id'], a['game_id'])] = a['count']
        
    for student in students:
        status_list = []
        for game in games:
            att_count = att_map.get((student.id, game.id), 0)
            allowed = perm_map.get((student.id, game.id), 1)
            is_blocked = att_count >= allowed
            status_list.append({
                'game': game,
                'is_blocked': is_blocked,
            })
        student.game_statuses = status_list
        
    return render(request, 'game_core/view_students.html', {
        'students': students,
        'search_query': query,
        'games': games,
    })

@login_required
def create_student(request):
    from django.contrib.auth.hashers import make_password
    import openpyxl
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'manual':
            name = request.POST.get('name', '').strip()
            register_number = request.POST.get('register_number', '').strip()
            email = request.POST.get('email', '').strip()
            
            if not name or not register_number:
                messages.error(request, "Please provide both Name and Register Number.")
            else:
                if Student.objects.filter(register_number=register_number).exists():
                    messages.error(request, f"A student with Register Number '{register_number}' already exists.")
                else:
                    default_password = f"{register_number}@{register_number}"
                    Student.objects.create(
                        name=name,
                        register_number=register_number,
                        email=email,
                        password=make_password(default_password)
                    )
                    messages.success(request, f"Student '{name}' created successfully with password '{default_password}'.")
                    return redirect('view_students')
                    
        elif action == 'excel':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, "Please select an Excel file to upload.")
            else:
                try:
                    wb = openpyxl.load_workbook(excel_file, read_only=True)
                    sheet = wb.active
                    
                    success_count = 0
                    skip_count = 0
                    
                    header_row = None
                    for row in sheet.iter_rows(values_only=True):
                        if any(row):  
                            header_row = row
                            break
                    
                    name_idx = 0
                    reg_idx = 1
                    email_idx = -1
                    
                    if header_row:
                        header_lower = [str(cell).lower().strip() if cell is not None else "" for cell in header_row]
                        for idx, val in enumerate(header_lower):
                            if "name" in val:
                                name_idx = idx
                            elif "register" in val or "reg" in val or "number" in val or "roll" in val:
                                reg_idx = idx
                            elif "email" in val or "mail" in val:
                                email_idx = idx
                                
                    first_row_skipped = False
                    for row in sheet.iter_rows(values_only=True):
                        if not any(row):
                            continue
                        
                        if not first_row_skipped and header_row and row == header_row:
                            first_row_skipped = True
                            continue
                            
                        name = str(row[name_idx]).strip() if len(row) > name_idx and row[name_idx] is not None else None
                        register_number = str(row[reg_idx]).strip() if len(row) > reg_idx and row[reg_idx] is not None else None
                        email = str(row[email_idx]).strip() if (email_idx != -1 and len(row) > email_idx and row[email_idx] is not None) else ''
                        
                        if not name or not register_number or name.lower() == 'none' or register_number.lower() == 'none':
                            continue
                        
                        if Student.objects.filter(register_number=register_number).exists():
                            skip_count += 1
                        else:
                            default_password = f"{register_number}@{register_number}"
                            Student.objects.create(
                                name=name,
                                register_number=register_number,
                                email=email,
                                password=make_password(default_password)
                            )
                            success_count += 1
                            
                    messages.success(request, f"Imported {success_count} students successfully. Skipped {skip_count} duplicates/invalid rows.")
                    return redirect('view_students')
                except Exception as e:
                    messages.error(request, f"Failed to process Excel file: {str(e)}")
                    
    return render(request, 'game_core/create_student.html')

@login_required
def edit_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        register_number = request.POST.get('register_number', '').strip()
        email = request.POST.get('email', '').strip()
        password_raw = request.POST.get('password', '').strip()
        
        if not name or not register_number:
            messages.error(request, "Name and Register Number are required.")
        else:
            if Student.objects.filter(register_number=register_number).exclude(id=student.id).exists():
                messages.error(request, f"Register number '{register_number}' is already taken.")
            else:
                student.name = name
                student.register_number = register_number
                student.email = email
                if password_raw:
                    from django.contrib.auth.hashers import make_password
                    student.password = make_password(password_raw)
                    messages.info(request, "Student password updated successfully.")
                student.save()
                messages.success(request, "Student details updated successfully.")
                return redirect('view_students')
                
    return render(request, 'game_core/edit_student.html', {'student': student})

@login_required
def student_profile(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    
    # Load attempts
    attempts = GameAttempt.objects.filter(student=student).select_related('game').order_by('-completed_at')
    
    # Load marks
    marks = StudentMark.objects.filter(student=student).select_related('subject').order_by('subject__name', 'exam_type')
    
    # Calculate average mark
    avg_marks = marks.aggregate(Avg('marks'))['marks__avg'] or 0
    avg_marks = round(avg_marks, 1)
    
    return render(request, 'game_core/student_profile.html', {
        'student': student,
        'attempts': attempts,
        'marks': marks,
        'avg_marks': avg_marks,
    })

@login_required
def manage_announcements(request):
    announcements = Announcement.objects.filter(created_by=request.user).order_by('-created_at')
    feedback_list = request.user.received_feedback.all().order_by('-created_at')
    teacher_settings, _ = TeacherSetting.objects.get_or_create(teacher=request.user)
    
    if request.method == 'POST':
        message = request.POST.get('message', '').strip()
        
        if not message:
            messages.error(request, "Please type a message to broadcast.")
        else:
            Announcement.objects.create(
                message=message,
                created_by=request.user
            )
            messages.success(request, "Announcement broadcasted successfully!")
            return redirect('manage_announcements')
            
    return render(request, 'game_core/manage_announcements.html', {
        'announcements': announcements,
        'feedback_list': feedback_list,
        'teacher_settings': teacher_settings,
    })

@login_required
def delete_announcement(request, announcement_id):
    announcement = get_object_or_404(Announcement, id=announcement_id, created_by=request.user)
    announcement.delete()
    messages.success(request, "Announcement deleted successfully.")
    return redirect('manage_announcements')

@login_required
def delete_feedback(request, feedback_id):
    feedback = get_object_or_404(AnonymousFeedback, id=feedback_id, teacher=request.user)
    feedback.delete()
    messages.success(request, "Anonymous message deleted successfully.")
    return redirect('manage_announcements')

@login_required
def toggle_feedback_read(request, feedback_id):
    feedback = get_object_or_404(AnonymousFeedback, id=feedback_id, teacher=request.user)
    feedback.is_read = not feedback.is_read
    feedback.save()
    status = "read" if feedback.is_read else "unread"
    messages.success(request, f"Message marked as {status}.")
    return redirect('manage_announcements')

# Student Game Creation & Teacher Approval Views
@login_required
def toggle_student_game_creation(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    student.can_create_games = not student.can_create_games
    student.save()
    status = "allowed" if student.can_create_games else "removed"
    messages.success(request, f"Game creation access for student '{student.name}' is now {status}.")
    return redirect('view_students')

@student_login_required
def student_create_game_setup(request):
    student_id = request.session.get('student_id')
    student = get_object_or_404(Student, id=student_id)
    
    if not student.can_create_games:
        messages.error(request, "You do not have permission to create games.")
        return redirect('student_dashboard')
        
    if request.method == 'POST':
        title = request.POST.get('title')
        game_type = request.POST.get('game_type')
        instructions = request.POST.get('instructions')
        explanation = request.POST.get('explanation')
        
        subject, _ = Subject.objects.get_or_create(name="General")
        
        game = Game.objects.create(
            title=title,
            subject=subject,
            game_type=game_type,
            instructions=instructions,
            explanation=explanation,
            created_by_student=student,
            approval_status='draft',
            is_active=False
        )
        return redirect('student_setup_game_questions', game_id=game.id)
        
    return render(request, 'game_core/student_create_game.html')

@student_login_required
def student_setup_game_questions(request, game_id):
    student_id = request.session.get('student_id')
    student = get_object_or_404(Student, id=student_id)
    game = get_object_or_404(Game, id=game_id, created_by_student=student)
    
    if not student.can_create_games:
        messages.error(request, "You do not have permission to create games.")
        return redirect('student_dashboard')
        
    if request.method == 'POST':
        if game.game_type == 'word_puzzle':
            from .models import WordPuzzleQuestion
            game.word_puzzles.all().delete()
            words = request.POST.getlist('correct_word')
            overrides = request.POST.getlist('letters_override')
            c1s = request.POST.getlist('clue1')
            c1ts = request.POST.getlist('clue1_time')
            c2s = request.POST.getlist('clue2')
            c2ts = request.POST.getlist('clue2_time')
            c3s = request.POST.getlist('clue3')
            fms = request.POST.getlist('full_marks')
            r1s = request.POST.getlist('reduced_marks_clue1')
            m2s = request.POST.getlist('min_marks_clue2')
            
            for i in range(len(words)):
                if words[i].strip():
                    WordPuzzleQuestion.objects.create(
                        game=game,
                        correct_word=words[i].strip().upper(),
                        letters_override=overrides[i].strip().upper() if i < len(overrides) else '',
                        clue1=c1s[i],
                        clue1_time=int(c1ts[i] or 15),
                        clue2=c2s[i],
                        clue2_time=int(c2ts[i] or 30),
                        clue3=c3s[i] if i < len(c3s) else '',
                        full_marks=int(fms[i] or 10),
                        reduced_marks_clue1=int(r1s[i] or 7),
                        min_marks_clue2=int(m2s[i] or 4)
                    )
            messages.success(request, f"Word Puzzle Game '{game.title}' questions saved successfully!")
            
        elif game.game_type == 'mcq':
            game.mcq_questions.all().delete()
            q_texts = request.POST.getlist('q_text')
            opt_as = request.POST.getlist('opt_a')
            opt_bs = request.POST.getlist('opt_b')
            opt_cs = request.POST.getlist('opt_c')
            opt_ds = request.POST.getlist('opt_d')
            corrects = request.POST.getlist('correct')
            c1s = request.POST.getlist('clue1')
            c2s = request.POST.getlist('clue2')
            c3s = request.POST.getlist('clue3')
            
            for i in range(len(q_texts)):
                if q_texts[i].strip():
                    MCQQuestion.objects.create(
                        game=game,
                        question_text=q_texts[i],
                        option_a=opt_as[i],
                        option_b=opt_bs[i],
                        option_c=opt_cs[i],
                        option_d=opt_ds[i],
                        correct_option=corrects[i],
                        clue1=c1s[i] if i < len(c1s) else '',
                        clue2=c2s[i] if i < len(c2s) else '',
                        clue3=c3s[i] if i < len(c3s) else ''
                    )
            messages.success(request, f"MCQ Game '{game.title}' questions saved successfully!")
            
        elif game.game_type == 'fill_blanks':
            game.blank_questions.all().delete()
            q_texts = request.POST.getlist('q_text')
            answers = request.POST.getlist('correct_answer')
            c1s = request.POST.getlist('clue1')
            c2s = request.POST.getlist('clue2')
            c3s = request.POST.getlist('clue3')
            
            for i in range(len(q_texts)):
                if q_texts[i].strip():
                    FillBlankQuestion.objects.create(
                        game=game,
                        question_text=q_texts[i],
                        correct_answer=answers[i].strip(),
                        clue=c1s[i].strip() if i < len(c1s) else '',
                        clue1=c1s[i].strip() if i < len(c1s) else '',
                        clue2=c2s[i].strip() if i < len(c2s) else '',
                        clue3=c3s[i].strip() if i < len(c3s) else ''
                    )
            messages.success(request, f"Fill in the Blanks Game '{game.title}' questions saved successfully!")
            
        elif game.game_type == 'match_following':
            game.match_items.all().delete()
            lefts = request.POST.getlist('left_item')
            rights = request.POST.getlist('right_item')
            
            game.clue1 = request.POST.get('clue1', '').strip()
            game.clue2 = request.POST.get('clue2', '').strip()
            game.clue3 = request.POST.get('clue3', '').strip()
            game.save()
            
            for i in range(len(lefts)):
                if lefts[i].strip() and rights[i].strip():
                    MatchItem.objects.create(
                        game=game,
                        left_item=lefts[i].strip(),
                        right_item=rights[i].strip()
                    )
            messages.success(request, f"Match the Following Game '{game.title}' questions saved successfully!")
            
        return redirect('student_dashboard')
        
    return render(request, f'game_core/setup_{game.game_type}.html', {
        'game': game,
        'is_student': True,
    })

@student_login_required
def student_submit_game(request, game_id):
    student_id = request.session.get('student_id')
    student = get_object_or_404(Student, id=student_id)
    game = get_object_or_404(Game, id=game_id, created_by_student=student)
    
    if not student.can_create_games:
        messages.error(request, "You do not have permission to create games.")
        return redirect('student_dashboard')
        
    game.approval_status = 'pending'
    game.save()
    messages.success(request, f"Game '{game.title}' has been submitted to your teacher for review.")
    return redirect('student_dashboard')

@login_required
def review_student_games(request):
    student_games = Game.objects.filter(created_by_student__isnull=False).select_related('created_by_student').order_by('-created_at')
    
    pending_games = student_games.filter(approval_status='pending')
    approved_games = student_games.filter(approval_status='approved')
    rejected_games = student_games.filter(approval_status='rejected')
    
    return render(request, 'game_core/review_student_games.html', {
        'pending_games': pending_games,
        'approved_games': approved_games,
        'rejected_games': rejected_games,
    })

@login_required
def approve_student_game(request, game_id):
    game = get_object_or_404(Game, id=game_id, created_by_student__isnull=False)
    game.approval_status = 'approved'
    game.is_active = True
    game.rejection_reason = None
    game.save()
    messages.success(request, f"Student game '{game.title}' has been Approved and Published online!")
    return redirect('review_student_games')

@login_required
def reject_student_game(request, game_id):
    game = get_object_or_404(Game, id=game_id, created_by_student__isnull=False)
    if request.method == 'POST':
        reason = request.POST.get('rejection_reason', '').strip()
        game.approval_status = 'rejected'
        game.rejection_reason = reason
        game.is_active = False
        game.save()
        messages.success(request, f"Student game '{game.title}' has been Rejected.")
    return redirect('review_student_games')

@student_login_required
def student_view_games(request):
    student = get_object_or_404(Student, id=request.session['student_id'])
    
    # Load settings for the teacher
    teacher = User.objects.filter(username='teacher').first()
    teacher_settings = None
    if teacher:
        teacher_settings, _ = TeacherSetting.objects.get_or_create(teacher=teacher)
    else:
        first_setting = TeacherSetting.objects.first()
        if first_setting:
            teacher_settings = first_setting
        else:
            teacher_settings = TeacherSetting()
            
    # Filtering for games
    type_filter = request.GET.get('type')
    
    games = Game.objects.filter(is_active=True, approval_status='approved').prefetch_related('word_puzzles', 'mcq_questions', 'blank_questions', 'match_items')
    if type_filter:
        games = games.filter(game_type=type_filter)
        
    games = games.order_by('-created_at')

    filtered_games = []
    for g in games:
        if g.game_type == 'word_puzzle' and len(g.word_puzzles.all()) > 0:
            filtered_games.append(g)
        elif g.game_type == 'mcq' and len(g.mcq_questions.all()) > 0:
            filtered_games.append(g)
        elif g.game_type == 'fill_blanks' and len(g.blank_questions.all()) > 0:
            filtered_games.append(g)
        elif g.game_type == 'match_following' and len(g.match_items.all()) > 0:
            filtered_games.append(g)
            
    # Apply global games enabled option
    if not teacher_settings.games_enabled:
        games_list = []
    else:
        games_list = filtered_games
        
    # Build attempts and block states map
    student_attempts = GameAttempt.objects.filter(student=student).values('game_id').annotate(count=Count('id'))
    attempts_map = {item['game_id']: item['count'] for item in student_attempts}
    
    from .models import StudentGamePlayPermission
    permissions = StudentGamePlayPermission.objects.filter(student=student)
    permissions_map = {perm.game_id: perm.allowed_attempts for perm in permissions}

    blocked_game_ids = []
    for g in games_list:
        attempts_count = attempts_map.get(g.id, 0)
        allowed = permissions_map.get(g.id, 1)
        g.student_attempts_count = attempts_count
        g.allowed_attempts = allowed
        if not teacher_settings.replays_enabled and attempts_count >= 1:
            blocked_game_ids.append(g.id)
        elif attempts_count >= allowed:
            blocked_game_ids.append(g.id)
            
    return render(request, 'game_core/student_view_games.html', {
        'games': games_list,
        'selected_type': type_filter,
        'attempts_map': attempts_map,
        'permissions_map': permissions_map,
        'blocked_game_ids': blocked_game_ids,
        'teacher_settings': teacher_settings,
    })

@student_login_required
def student_view_marks(request):
    student = get_object_or_404(Student, id=request.session['student_id'])
    
    # Fetch student marks
    my_marks = StudentMark.objects.filter(student=student).select_related('subject').order_by('subject__name', 'exam_type')
    
    # Aggregate exams
    all_exams = list(my_marks.values_list('exam_type', flat=True).distinct())
    
    # Simple performance comparison between Exam 1 and Exam 2
    improved_list = []
    declined_list = []
    stable_list = []
    exam1_scores = {}
    exam2_scores = {}
    
    for mark in my_marks:
        # Calculate percentage
        percentage = (mark.marks / mark.max_marks) * 100 if mark.max_marks > 0 else 0
        mark.percentage = percentage
        if mark.exam_type == 'Exam 1':
            exam1_scores[mark.subject.name] = percentage
        elif mark.exam_type == 'Exam 2':
            exam2_scores[mark.subject.name] = percentage
            
    for sub_name, score2 in exam2_scores.items():
        if sub_name in exam1_scores:
            score1 = exam1_scores[sub_name]
            if score2 > score1:
                improved_list.append(sub_name)
            elif score2 < score1:
                declined_list.append(sub_name)
            else:
                stable_list.append(sub_name)
                
    eval_summary = ""
    eval_status = ""
    if my_marks.exists():
        if 'Exam 1' in all_exams or 'Exam 2' in all_exams:
            parts = []
            if improved_list:
                parts.append(f"improved in {', '.join(improved_list)}")
            if declined_list:
                parts.append(f"declined in {', '.join(declined_list)}")
            if stable_list:
                parts.append(f"remained stable in {', '.join(stable_list)}")
                
            if parts:
                eval_summary = f"Compared to Exam 1, you {'; and you '.join(parts)}."
            else:
                eval_summary = "Once marks for both Exam 1 and Exam 2 are uploaded, your detailed subject comparison will appear here."
                
            total_improved = len(improved_list)
            total_declined = len(declined_list)
            if total_improved > total_declined:
                eval_status = "improved"
            elif total_declined > total_improved:
                eval_status = "declined"
            else:
                eval_status = "stable"
        else:
            eval_summary = "Academic marks have been uploaded for other examinations. Auto-comparison runs on 'Exam 1' vs 'Exam 2'."
            eval_status = "other_exams"
    else:
        eval_summary = "No marks records have been posted yet."
        eval_status = "no_data"
        
    return render(request, 'game_core/student_view_marks.html', {
        'my_marks': my_marks,
        'eval_summary': eval_summary,
        'eval_status': eval_status,
    })

@student_login_required
def student_view_notes(request):
    notes = PDFNote.objects.all().select_related('subject', 'uploaded_by').order_by('-uploaded_at')
    return render(request, 'game_core/student_view_notes.html', {
        'notes': notes,
    })

@student_login_required
def student_send_feedback(request):
    student = get_object_or_404(Student, id=request.session['student_id'])
    
    # Fetch settings to see if enabled
    teacher = User.objects.filter(username='teacher').first()
    teacher_settings = None
    if teacher:
        teacher_settings, _ = TeacherSetting.objects.get_or_create(teacher=teacher)
    else:
        teacher_settings = TeacherSetting()
        
    if request.method == 'POST':
        message = request.POST.get('message', '').strip()
        if not message:
            messages.error(request, "Please enter a message suggestions/feedback.")
        else:
            teachers = User.objects.all()
            if teachers.exists():
                for t in teachers:
                    AnonymousFeedback.objects.create(
                        teacher=t,
                        message=message
                    )
                messages.success(request, "Your anonymous suggestion has been sent successfully!")
            else:
                messages.error(request, "No teacher exists to receive suggestions.")
            return redirect('student_dashboard')
            
    return render(request, 'game_core/student_send_feedback.html', {
        'teacher_settings': teacher_settings,
    })

@student_login_required
def student_performance_analysis(request):
    student = get_object_or_404(Student, id=request.session['student_id'])
    
    my_attempts = GameAttempt.objects.filter(student=student).order_by('completed_at')
    progress_labels = [att.completed_at.strftime('%m/%d %H:%M') for att in my_attempts]
    progress_scores = [att.score for att in my_attempts]
    
    leaderboard = GameAttempt.objects.values(
        'student__name', 'student__register_number'
    ).annotate(
        total_score=Avg('score'),
        avg_time=Avg('time_taken'),
        games_played=Count('game', distinct=True)
    ).order_by('-total_score', 'avg_time')[:10]
    
    my_marks = StudentMark.objects.filter(student=student).select_related('subject').order_by('subject__name', 'exam_type')
    
    comparison_subjects = []
    exam1_scores = []
    exam2_scores = []
    
    e1_map = {}
    e2_map = {}
    subjects = set()
    for m in my_marks:
        percentage = (m.marks / m.max_marks) * 100 if m.max_marks > 0 else 0
        if m.exam_type == 'Exam 1':
            e1_map[m.subject.name] = percentage
            subjects.add(m.subject.name)
        elif m.exam_type == 'Exam 2':
            e2_map[m.subject.name] = percentage
            subjects.add(m.subject.name)
            
    for sub in sorted(list(subjects)):
        comparison_subjects.append(sub)
        exam1_scores.append(e1_map.get(sub, 0))
        exam2_scores.append(e2_map.get(sub, 0))
        
    return render(request, 'game_core/student_performance_analysis.html', {
        'student': student,
        'my_attempts': my_attempts.order_by('-completed_at')[:10],
        'progress_labels': progress_labels,
        'progress_scores': progress_scores,
        'leaderboard': leaderboard,
        'comparison_subjects': comparison_subjects,
        'exam1_scores': exam1_scores,
        'exam2_scores': exam2_scores,
    })


@login_required
def game_insights(request, game_id):
    from django.db.models import Avg, Max, Min
    from .models import Game, GameEntry, GameAttempt
    
    game = get_object_or_404(Game, id=game_id)
    
    # 1. Player Tracking Metrics
    total_entries = GameEntry.objects.filter(game=game).count()
    # Unique students count (Total Players)
    unique_players_count = GameEntry.objects.filter(game=game).values('student').distinct().count()
    
    # 2. Performance Summary Metrics
    attempts_qs = GameAttempt.objects.filter(game=game)
    total_attempts = attempts_qs.count()
    
    avg_score = attempts_qs.aggregate(avg=Avg('score'))['avg']
    high_score = attempts_qs.aggregate(max=Max('score'))['max']
    low_score = attempts_qs.aggregate(min=Min('score'))['min']
    avg_time = attempts_qs.aggregate(avg=Avg('time_taken'))['avg']
    
    # Format metrics
    avg_score = round(avg_score, 1) if avg_score is not None else 0
    high_score = high_score if high_score is not None else 0
    low_score = low_score if low_score is not None else 0
    avg_time = round(avg_time, 1) if avg_time is not None else 0
    
    completion_rate = 0
    if total_entries > 0:
        completion_rate = round((total_attempts / total_entries) * 100, 1)
        
    # 3. Leaderboard calculation
    leaderboard = get_game_leaderboard(game, limit=None)  # Full leaderboard
    
    # 4. Participants List
    entered_student_ids = GameEntry.objects.filter(game=game).values_list('student_id', flat=True).distinct()
    entered_students = Student.objects.filter(id__in=entered_student_ids)
    
    participants = []
    for student in entered_students:
        first_entry = GameEntry.objects.filter(game=game, student=student).order_by('entry_time').first()
        best_attempt = GameAttempt.objects.filter(game=game, student=student).order_by('-score', 'time_taken').first()
        attempts_count = GameAttempt.objects.filter(game=game, student=student).count()
        
        participants.append({
            'student': student,
            'first_entered': first_entry.entry_time if first_entry else None,
            'best_attempt': best_attempt,
            'attempts_count': attempts_count,
            'completed': attempts_count > 0
        })
        
    # Sort participants: completed first, then by score descending, then by name
    def sort_key(p):
        best_score = p['best_attempt'].score if p['best_attempt'] else -1
        best_time = p['best_attempt'].time_taken if p['best_attempt'] else 999999
        return (-int(p['completed']), -best_score, best_time, p['student'].name.lower())
        
    participants.sort(key=sort_key)
    
    context = {
        'game': game,
        'total_entries': total_entries,
        'unique_players_count': unique_players_count,
        'total_attempts': total_attempts,
        'avg_score': avg_score,
        'high_score': high_score,
        'low_score': low_score,
        'avg_time': avg_time,
        'completion_rate': completion_rate,
        'leaderboard': leaderboard[:10],
        'full_leaderboard': leaderboard,
        'participants': participants,
    }
    
    return render(request, 'game_core/game_insights.html', context)


@student_login_required
def student_change_password(request):
    student = get_object_or_404(Student, id=request.session['student_id'])
    
    if request.method == 'POST':
        current_password = request.POST.get('current_password', '')
        new_password = request.POST.get('new_password', '')
        confirm_password = request.POST.get('confirm_password', '')
        
        from django.contrib.auth.hashers import check_password, make_password
        if not student.password or not check_password(current_password, student.password):
            messages.error(request, "Incorrect current password.")
            return render(request, 'game_core/student_change_password.html')
            
        if new_password != confirm_password:
            messages.error(request, "New password and confirmation password do not match.")
            return render(request, 'game_core/student_change_password.html')
            
        if not new_password.strip():
            messages.error(request, "New password cannot be empty.")
            return render(request, 'game_core/student_change_password.html')
            
        student.password = make_password(new_password)
        student.save()
        
        request.session.flush()
        messages.success(request, "Password updated successfully. Please login again with your new credentials.")
        return redirect('student_login')
        
    return render(request, 'game_core/student_change_password.html')


